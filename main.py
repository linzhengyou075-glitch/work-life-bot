from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo
from io import BytesIO
from pathlib import Path
import asyncio
import logging
import sqlite3
import time
import re
import urllib.parse

from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, FileResponse
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from openpyxl import load_workbook

from auth import build_authorize_url, exchange_code, get_profile, new_state, LoginConfigError
from config import settings
from database import (
    init_db,upsert_user,list_shift_types,list_shifts,get_shift,save_shift,delete_shift,
    list_daily_work,toggle_daily_work,list_logistics_settings,save_logistics_setting,
    delete_logistics_setting,list_daily_logistics,logistics_arrive,logistics_complete,mark_logistics_reminded,
    attendance_for,check_in,check_out,add_reminder,list_reminders,due_reminders,
    mark_reminder_sent,complete_reminder,delete_reminder,add_notification_log,
    list_notification_logs,add_task,list_tasks,toggle_task,delete_task,
    add_activity,get_activity,update_activity,complete_activity,delete_activity,mark_activity_reminded,due_activities,list_activities,upcoming_activities,
    add_work_log,list_work_logs,delete_work_log,dashboard_counts,set_setting,get_setting,
    database_status,database_health_check,run_daily_maintenance,latest_maintenance_log
)
from line_api import verify_signature,reply_message,work_entry_flex,push_message,reminder_flex

BASE_DIR = Path(__file__).resolve().parent
TAIPEI_TZ = ZoneInfo("Asia/Taipei")
_SCHEDULER_TASK = None
_SCHEDULER_STOP_EVENT = None
_SCHEDULER_INTERVAL_SECONDS = 60
_STATIC_CACHE_SECONDS = 300
_IMAGE_CACHE_SECONDS = 86400
_REMINDER_WORKER_STATE = {
    "running": False,
    "last_check": None,
    "last_error": None,
    "task_name": "work-life-scheduler",
    "interval_seconds": _SCHEDULER_INTERVAL_SECONDS,
    "last_maintenance": None,
    "cycle_count": 0,
    "last_cycle_ms": None,
    "last_push_retry_count": 0,
}

logger = logging.getLogger("worklife")

async def _push_with_retry(user_id, messages, attempts=3):
    """LINE 推播短暫失敗時最多重試 3 次，且不阻塞 FastAPI 事件迴圈。"""
    max_attempts = max(1, int(attempts or 1))
    last_error = None
    for attempt in range(1, max_attempts + 1):
        try:
            # push_message 是同步網路呼叫，移至工作執行緒避免卡住網站主迴圈。
            await asyncio.to_thread(push_message, user_id, messages)
            _REMINDER_WORKER_STATE["last_push_retry_count"] = attempt - 1
            return
        except Exception as exc:
            last_error = exc
            if attempt < max_attempts:
                await asyncio.sleep(min(2 ** (attempt - 1), 4))
    _REMINDER_WORKER_STATE["last_push_retry_count"] = max(0, max_attempts - 1)
    if last_error is not None:
        raise last_error
    raise RuntimeError("LINE 推播失敗，但未取得錯誤原因")


def _time_minutes(value):
    try:
        h,m=str(value or "").split(":")[:2]
        return int(h)*60+int(m)
    except Exception:
        return 0

def _is_overnight_shift(shift):
    return bool(shift) and _time_minutes(shift.get("end_time")) <= _time_minutes(shift.get("start_time"))

def _shift_end_value(shift):
    return (shift.get("overtime_end") if shift and shift.get("overtime") else shift.get("end_time")) if shift else ""

def _display_md(value):
    return value.strftime("%m/%d")

def _shift_display_name(shift):
    """組合班表通知顯示名稱，兼容不同資料庫欄位版本。"""
    if not shift:
        return "上班"
    for key in ("shift_name", "name", "shift_type_name"):
        value=str(shift.get(key) or "").strip()
        if value:
            return value
    store=str(shift.get("store_name") or shift.get("store_code") or "").strip()
    return f"{store}上班" if store else "上班"

def _shift_store_name(shift):
    if not shift:
        return ""
    value=str(shift.get("store_name") or "").strip()
    if value:
        return value
    code=str(shift.get("store_code") or "").upper()
    return {"B":"太平建昌店","C":"太平溪洲店"}.get(code,"")

def decorate_shift_dates(work_date, shift):
    if not shift:
        return None
    item=dict(shift)
    start_day=datetime.strptime(work_date,"%Y-%m-%d").date()
    end_day=start_day+timedelta(days=1) if _time_minutes(_shift_end_value(item)) <= _time_minutes(item.get("start_time")) else start_day
    item.update({
        "work_date":work_date,
        "start_date":start_day.isoformat(),
        "end_date":end_day.isoformat(),
        "start_display":f"{_display_md(start_day)} {item.get('start_time') or '--:--'}",
        "end_display":f"{_display_md(end_day)} {_shift_end_value(item) or '--:--'}",
        "range_display":f"{_display_md(start_day)} {item.get('start_time') or '--:--'} ～ {_display_md(end_day)} {_shift_end_value(item) or '--:--'}",
        "is_overnight":end_day!=start_day,
    })
    return item

def decorate_timed_rows(work_date, shift, rows, time_key="scheduled_time"):
    start_day=datetime.strptime(work_date,"%Y-%m-%d").date()
    overnight=_is_overnight_shift(shift)
    start_minutes=_time_minutes(shift.get("start_time")) if shift else 0
    output=[]
    for row in rows:
        item=dict(row)
        raw=item.get(time_key) or ""
        item_minutes=_time_minutes(raw)
        item_day=start_day+timedelta(days=1) if overnight and item_minutes < start_minutes else start_day
        item["display_date"]=item_day.isoformat()
        item["display_md"]=_display_md(item_day)
        item["datetime_display"]=f"{_display_md(item_day)} {raw or '--:--'}"
        item["sort_datetime"]=f"{item_day.isoformat()} {raw or '23:59'}"
        output.append(item)
    return sorted(output,key=lambda x:(x["sort_datetime"],x.get("id",0)))

def resolve_active_work_date(now=None):
    now=now or datetime.now()
    today=now.date()
    previous=today-timedelta(days=1)
    previous_shift=get_shift(previous.isoformat())
    if previous_shift and _is_overnight_shift(previous_shift):
        end_value=_shift_end_value(previous_shift)
        end_at=datetime.combine(today,datetime.strptime(end_value,"%H:%M").time())
        if now <= end_at:
            return previous.isoformat()
    return today.isoformat()

def dashboard_work_context(now=None):
    now=now or datetime.now()
    work_date=resolve_active_work_date(now)
    raw_shift=get_shift(work_date)
    shift=decorate_shift_dates(work_date,raw_shift)
    work=decorate_timed_rows(work_date,raw_shift,list_daily_work(work_date),"scheduled_time")
    logistics=decorate_timed_rows(work_date,raw_shift,list_daily_logistics(work_date),"start_time")
    next_item=None
    for item in work:
        try:
            at=datetime.strptime(f"{item['display_date']} {item.get('scheduled_time')}","%Y-%m-%d %H:%M")
        except Exception:
            continue
        if not item.get("is_done") and at>=now:
            next_item=item
            break
    return work_date,shift,work,logistics,next_item

def get_next_shift(after_date=None):
    after_date=after_date or date.today().isoformat()
    for item in list_shifts():
        if item.get("work_date", "") < after_date:
            continue
        if item.get("store_code") == "X":
            continue
        return decorate_shift_dates(item["work_date"], item)
    return None


def week_schedule(anchor=None):
    anchor=anchor or date.today()
    start=anchor-timedelta(days=anchor.weekday())
    labels="一二三四五六日"
    output=[]
    for i in range(7):
        day=start+timedelta(days=i)
        shift=get_shift(day.isoformat())
        output.append({"date":day.isoformat(),"day":day.day,"weekday":labels[i],"today":day==anchor,"shift":dict(shift) if shift else None})
    return output


app=FastAPI(title="Work Life",version="1.0.5.2")
app.add_middleware(
    SessionMiddleware,
    secret_key=settings.session_secret,
    max_age=60*60*24*30,
    same_site="lax",
    https_only=settings.base_url.startswith("https://")
)
templates=Jinja2Templates(directory=str(BASE_DIR))

def _static_response(filename, media_type, cache_seconds):
    """讓瀏覽器短期快取靜態檔，減少重複向 Render 下載。"""
    response=FileResponse(BASE_DIR / filename,media_type=media_type)
    response.headers["Cache-Control"]=f"public, max-age={cache_seconds}"
    return response

@app.get("/static/style.css")
def static_style(): return _static_response("style.css","text/css",_STATIC_CACHE_SECONDS)
@app.get("/static/app.js")
def static_js(): return _static_response("app.js","application/javascript",_STATIC_CACHE_SECONDS)
@app.get("/static/worklife_mascot.png")
def static_mascot(): return _static_response("worklife_mascot.png","image/png",_IMAGE_CACHE_SECONDS)
@app.get("/static/worklife_frame.png")
def static_frame(): return _static_response("worklife_frame.png","image/png",_IMAGE_CACHE_SECONDS)

@app.on_event("startup")
async def startup():
    """啟動唯一一個背景排程器。"""
    global _SCHEDULER_TASK, _SCHEDULER_STOP_EVENT
    init_db()

    # 同一個 Render 程序只允許存在一個排程工作，避免重複推播與資源浪費。
    if _SCHEDULER_TASK is not None and not _SCHEDULER_TASK.done():
        return

    _SCHEDULER_STOP_EVENT = asyncio.Event()
    _SCHEDULER_TASK = asyncio.create_task(
        reminder_worker(),
        name=_REMINDER_WORKER_STATE["task_name"],
    )

@app.on_event("shutdown")
async def shutdown():
    """Render 關閉或重新部署時，乾淨停止排程器。"""
    global _SCHEDULER_TASK, _SCHEDULER_STOP_EVENT
    if _SCHEDULER_STOP_EVENT is not None:
        _SCHEDULER_STOP_EVENT.set()

    task = _SCHEDULER_TASK
    if task is not None and not task.done():
        try:
            await asyncio.wait_for(task, timeout=5)
        except asyncio.TimeoutError:
            task.cancel()
            try:
                await task
            except asyncio.CancelledError:
                pass

    _SCHEDULER_TASK = None
    _SCHEDULER_STOP_EVENT = None
    _REMINDER_WORKER_STATE["running"] = False

async def reminder_worker():
    """唯一背景 Scheduler：集中處理班表、提醒、活動與物流通知。"""
    _REMINDER_WORKER_STATE["running"] = True
    try:
        while _SCHEDULER_STOP_EVENT is not None and not _SCHEDULER_STOP_EVENT.is_set():
            try:
                # Render 主機預設通常是 UTC；所有班表與提醒皆以台灣時間判斷。
                cycle_started = time.monotonic()
                now=datetime.now(TAIPEI_TZ).replace(tzinfo=None)
                _REMINDER_WORKER_STATE["last_check"] = now.strftime("%Y-%m-%d %H:%M:%S")
                _REMINDER_WORKER_STATE["last_error"] = None

                # 個人設定關閉 LINE 推播時不送出，但排程器仍保持運作。
                line_push_ready = bool(settings.line_channel_access_token and settings.owner_line_user_id)
                line_push_enabled = line_push_ready and get_setting("line_push", "1") == "1"

                # 每天只做一次輕量維護；不因每分鐘排程反覆清理資料庫。
                maintenance_key = now.strftime("%Y-%m-%d")
                if now.hour >= 4 and get_setting("last_daily_maintenance", "") != maintenance_key:
                    try:
                        maintenance_result=run_daily_maintenance(now.date())
                        _REMINDER_WORKER_STATE["last_maintenance"] = maintenance_result
                        set_setting("last_daily_maintenance", maintenance_key)
                    except Exception as exc:
                        _REMINDER_WORKER_STATE["last_error"] = f"每日維護失敗：{type(exc).__name__}: {exc}"

                # 未啟用或尚未設定 LINE 時，不掃描班表、活動與物流資料，降低空轉查詢。
                if line_push_enabled:
                    today_text=now.strftime("%Y-%m-%d")
                    time_text=now.strftime("%H:%M")
                    # 班表通知直接由已儲存班表產生：上傳 Excel 後無須另外建立提醒。
                    # 每分鐘檢查今天與明天的班表，上班前 10 分鐘推播一次。
                    for shift_date in (now.date(), now.date()+timedelta(days=1)):
                        work_date=shift_date.isoformat()
                        raw_shift=get_shift(work_date)
                        if not raw_shift or raw_shift.get("store_code")=="X":
                            continue
                        start_time=str(raw_shift.get("start_time") or "").strip()
                        try:
                            shift_at=datetime.strptime(f"{work_date} {start_time}","%Y-%m-%d %H:%M")
                        except Exception as exc:
                            add_notification_log("班表通知",f"{work_date} 上班時間格式錯誤：{exc}","排程失敗")
                            continue
                        remind_at=shift_at-timedelta(minutes=10)
                        # 只有進入實際通知時間窗才查已發送標記，避免每分鐘無效讀取設定表。
                        if remind_at <= now < shift_at+timedelta(minutes=5):
                            marker_key=f"shift_push_sent:{work_date}:{start_time}"
                            already_sent=get_setting(marker_key,"")=="1"
                            if already_sent:
                                continue
                            shift_name=_shift_display_name(raw_shift)
                            store_name=_shift_store_name(raw_shift)
                            title="⏰ 還有 10 分鐘就要上班囉！"
                            parts=[f"📅 {shift_at.strftime('%m/%d')}",f"🕒 {start_time}",f"💼 {shift_name}"]
                            if store_name and store_name not in shift_name:
                                parts.insert(2,f"📍 {store_name}")
                            content="\n".join(parts)
                            try:
                                url=f"{settings.base_url}/work-records?work_date={work_date}"
                                await _push_with_retry(settings.owner_line_user_id,[reminder_flex(title,content,url)])
                                set_setting(marker_key,"1")
                                add_notification_log(title,content,"已推播")
                            except Exception as exc:
                                error=f"{type(exc).__name__}: {exc}"
                                add_notification_log(title,error,"推播失敗")
                                _REMINDER_WORKER_STATE["last_error"] = error

                    for item in due_reminders(today_text,time_text):
                        title=item["title"]
                        content=item["note"] or f'{item["remind_time"]} 提醒'
                        try:
                            url=f'{settings.base_url}{item["related_url"] or "/dashboard"}'
                            await _push_with_retry(settings.owner_line_user_id,[reminder_flex(title,content,url)])
                            # 只有 LINE API 確認成功後才標記，避免失敗卻被當成已發送。
                            mark_reminder_sent(item["id"])
                            add_notification_log(title,content,"已推播")
                        except Exception as exc:
                            error=f"{type(exc).__name__}: {exc}"
                            add_notification_log(title,error,"推播失敗")
                            _REMINDER_WORKER_STATE["last_error"] = error

                    # 活動中心 LINE 提醒。未填活動時間時，預設於活動當日 09:00 提醒。
                    for item in due_activities(today_text,time_text):
                        title=f"📢 活動提醒｜{item['title']}"
                        parts=[f"📅 {item['activity_date']} {item.get('activity_time') or '09:00'}"]
                        if item.get("location"):
                            parts.append(f"📍 {item['location']}")
                        if item.get("note"):
                            parts.append(f"📝 {item['note']}")
                        content="\n".join(parts)
                        try:
                            url=f"{settings.base_url}/activities"
                            await _push_with_retry(settings.owner_line_user_id,[reminder_flex(title,content,url)])
                            mark_activity_reminded(item["id"])
                            add_notification_log(title,content,"已推播")
                        except Exception as exc:
                            error=f"{type(exc).__name__}: {exc}"
                            add_notification_log(title,error,"推播失敗")
                            _REMINDER_WORKER_STATE["last_error"] = error

                    # 依班表自動推播物流提醒；大夜班凌晨時段會套用前一天班表。
                    for work_date in ((now.date()-timedelta(days=1)).isoformat(), now.date().isoformat()):
                        raw_shift=get_shift(work_date)
                        if not raw_shift or raw_shift.get("store_code")=="X":
                            continue
                        for item in decorate_timed_rows(work_date,raw_shift,list_daily_logistics(work_date),"start_time"):
                            if item.get("reminded_at") or item.get("completed_at") or not item.get("line_push"):
                                continue
                            try:
                                event_at=datetime.strptime(f"{item['display_date']} {item['start_time']}","%Y-%m-%d %H:%M")
                                remind_at=event_at-timedelta(minutes=int(item.get("remind_minutes") or 10))
                            except Exception as exc:
                                add_notification_log(item.get("name") or "物流提醒",f"時間格式錯誤：{exc}","排程失敗")
                                continue
                            if remind_at <= now < event_at+timedelta(minutes=5):
                                title=f"{item.get('icon') or '🚚'} {item.get('name')}即將到店"
                                content=f"預計 {item['datetime_display']} 到店"
                                if item.get("content"):
                                    content += f"｜{item['content']}"
                                try:
                                    url=f"{settings.base_url}/work-records?work_date={work_date}"
                                    await _push_with_retry(settings.owner_line_user_id,[reminder_flex(title,content,url)])
                                    mark_logistics_reminded(item["id"])
                                    add_notification_log(title,content,"已推播")
                                except Exception as exc:
                                    error=f"{type(exc).__name__}: {exc}"
                                    add_notification_log(title,error,"推播失敗")
                                    _REMINDER_WORKER_STATE["last_error"] = error
            except asyncio.CancelledError:
                raise
            except Exception as exc:
                # 不再靜默忽略錯誤，健康檢查頁與 Render Log 都能看到原因。
                _REMINDER_WORKER_STATE["last_error"] = f"{type(exc).__name__}: {exc}"
                logger.exception("Work Life scheduler error: %s", _REMINDER_WORKER_STATE["last_error"])

            _REMINDER_WORKER_STATE["cycle_count"] += 1
            _REMINDER_WORKER_STATE["last_cycle_ms"] = max(0, int((time.monotonic() - cycle_started) * 1000))

            # 使用可中斷等待；重新部署時不必硬等 60 秒，並且不建立額外 Timer。
            try:
                await asyncio.wait_for(
                    _SCHEDULER_STOP_EVENT.wait(),
                    timeout=_SCHEDULER_INTERVAL_SECONDS,
                )
            except asyncio.TimeoutError:
                pass
    finally:
        _REMINDER_WORKER_STATE["running"] = False

def current_user(request):
    return request.session.get("user")

def protected(request,path):
    user=current_user(request)
    if not user:
        request.session["login_next"]=path
        return None,RedirectResponse("/login-page",302)
    if settings.owner_line_user_id and user.get("userId")!=settings.owner_line_user_id:
        return None,templates.TemplateResponse("denied.html",{"request":request},status_code=403)
    return user,None

@app.get("/healthz")
def healthz():
    """輕量健康檢查：只在開啟此網址時執行，不新增背景程序。"""
    now = datetime.now(TAIPEI_TZ).replace(tzinfo=None)
    task_running = bool(_SCHEDULER_TASK and not _SCHEDULER_TASK.done())
    worker_running = task_running and _REMINDER_WORKER_STATE["running"]

    last_check_text = _REMINDER_WORKER_STATE.get("last_check")
    scheduler_stale = True
    if last_check_text:
        try:
            last_check_at = datetime.strptime(last_check_text, "%Y-%m-%d %H:%M:%S")
            scheduler_stale = (now - last_check_at).total_seconds() > (_SCHEDULER_INTERVAL_SECONDS * 3)
        except ValueError:
            scheduler_stale = True

    line_push_ready = bool(settings.line_channel_access_token and settings.owner_line_user_id)
    line_push_enabled = get_setting("line_push", "1") == "1"
    db_health = database_health_check()
    last_maintenance = _REMINDER_WORKER_STATE.get("last_maintenance") or latest_maintenance_log()
    maintenance_today = bool(last_maintenance and last_maintenance.get("run_date") == now.date().isoformat())
    maintenance_due = now.hour >= 4 and not maintenance_today

    problems = []
    if not worker_running:
        problems.append("scheduler_not_running")
    elif scheduler_stale:
        problems.append("scheduler_check_stale")
    if not db_health.get("ok"):
        problems.append("database_unavailable")
    if line_push_enabled and not line_push_ready:
        problems.append("line_push_config_missing")
    if maintenance_due:
        problems.append("daily_maintenance_pending")
    if _REMINDER_WORKER_STATE.get("last_error"):
        problems.append("scheduler_last_error")

    return {
        "status":"ok" if not problems else "warning",
        "version":"1.0.5.2",
        "checked_at":now.strftime("%Y-%m-%d %H:%M:%S"),
        "timezone":"Asia/Taipei",
        "problems":problems,
        "line":{
            "login_ready":settings.line_login_ready,
            "push_ready":line_push_ready,
            "push_enabled":line_push_enabled,
            "configuration_ok":(not line_push_enabled) or line_push_ready,
        },
        "database":{**database_status(), **db_health},
        "scheduler":{
            "running":worker_running,
            "stale":scheduler_stale,
            "last_check":last_check_text,
            "last_error":_REMINDER_WORKER_STATE.get("last_error"),
            "task_name":_REMINDER_WORKER_STATE["task_name"],
            "interval_seconds":_REMINDER_WORKER_STATE["interval_seconds"],
            "single_worker":True,
            "idle_scan_skipped_when_line_unavailable":True,
            "cycle_count":_REMINDER_WORKER_STATE.get("cycle_count",0),
            "last_cycle_ms":_REMINDER_WORKER_STATE.get("last_cycle_ms"),
            "last_push_retry_count":_REMINDER_WORKER_STATE.get("last_push_retry_count",0),
        },
        "daily_maintenance":{
            "enabled":True,
            "scheduled_time":"04:00 Asia/Taipei",
            "ran_today":maintenance_today,
            "pending":maintenance_due,
            "last_result":last_maintenance,
        },
    }

@app.head("/", include_in_schema=False)
def root_head():
    return HTMLResponse(status_code=200)

@app.get("/")
def root(request:Request):
    return RedirectResponse("/dashboard" if current_user(request) else "/login-page",302)

@app.get("/login-page",response_class=HTMLResponse)
def login_page(request:Request):
    if current_user(request): return RedirectResponse("/dashboard",302)
    return templates.TemplateResponse("login.html",{"request":request})

@app.get("/login")
def login(request:Request):
    try:
        state=new_state()
        request.session["oauth_state"]=state
        return RedirectResponse(build_authorize_url(state),302)
    except LoginConfigError as exc:
        return templates.TemplateResponse("error.html", {"request":request,"message":str(exc)+" 請在部署環境設定 LINE_LOGIN_CHANNEL_ID、LINE_LOGIN_CHANNEL_SECRET、BASE_URL。"}, status_code=503)

@app.get("/auth/line/callback")
def callback(request:Request,code:str|None=None,state:str|None=None,error:str|None=None):
    if error:
        return templates.TemplateResponse("error.html",{"request":request,"message":"LINE 登入失敗，請重新登入。"},status_code=400)
    expected=request.session.pop("oauth_state",None)
    if not code or not state or state!=expected:
        return templates.TemplateResponse("error.html",{"request":request,"message":"登入驗證失敗，請重新登入。"},status_code=400)
    try:
        token=exchange_code(code)
        profile=get_profile(token["access_token"])
    except Exception as exc:
        return templates.TemplateResponse("error.html", {"request":request,"message":f"LINE 登入連線失敗：{exc}"}, status_code=400)
    if settings.owner_line_user_id and profile.get("userId")!=settings.owner_line_user_id:
        return templates.TemplateResponse("denied.html",{"request":request},status_code=403)
    upsert_user(profile["userId"],profile.get("displayName","佑佑"),profile.get("pictureUrl"))
    request.session["user"]=profile
    next_url=request.session.pop("login_next","/dashboard")
    if not isinstance(next_url,str) or not next_url.startswith("/") or next_url.startswith("//"):
        next_url="/dashboard"
    return RedirectResponse(next_url,302)

@app.get("/logout")
def logout(request:Request):
    request.session.clear()
    return RedirectResponse("/login-page",302)

@app.get("/dashboard",response_class=HTMLResponse)
def dashboard(request:Request):
    user,resp=protected(request,"/dashboard")
    if resp:return resp
    now=datetime.now()
    work_date,shift,daily_work,daily_logistics,next_work_item=dashboard_work_context(now)
    today=date.today().isoformat()
    return templates.TemplateResponse("dashboard.html",{
        "request":request,"user":user,"today":today,"active_work_date":work_date,
        "shift":shift,
        "daily_work":daily_work,
        "daily_logistics":daily_logistics,
        "next_work_item":next_work_item,
        "attendance":attendance_for(work_date),
        "tasks":list_tasks(False)[:5],
        "reminders":[x for x in list_reminders(False) if x["remind_date"]==today],
        "counts":dashboard_counts(work_date),
        "activities":upcoming_activities(today,5),
        "week_days":week_schedule(),
        "next_shift":get_next_shift(today),
        "now_text":datetime.now().strftime("%Y/%m/%d %H:%M"),
        "today_label":datetime.now().strftime("%m月%d日"),

    })

@app.get("/quick-schedule",response_class=HTMLResponse)
@app.get("/schedule",response_class=HTMLResponse)
def schedule_page(request:Request):
    user,resp=protected(request,request.url.path)
    if resp:return resp
    return templates.TemplateResponse("schedule.html",{
        "request":request,"user":user,
        "shifts":list_shifts(),"shift_types":list_shift_types()
    })

@app.post("/schedule")
def schedule_save(
    request:Request,work_date:str=Form(...),shift_type_id:int=Form(...),
    overtime:str=Form(""),overtime_end:str=Form(""),
    manager_tasks:str=Form(""),note:str=Form("")
):
    user,resp=protected(request,"/quick-schedule")
    if resp:return resp
    try:
        save_shift(work_date,shift_type_id,overtime=="1",overtime_end,manager_tasks=="1",note.strip())
    except (ValueError, OSError) as exc:
        return templates.TemplateResponse("schedule.html",{
            "request":request,"user":user,"shifts":list_shifts(),"shift_types":list_shift_types(),
            "save_error":str(exc)
        },status_code=400)
    except Exception:
        return templates.TemplateResponse("schedule.html",{
            "request":request,"user":user,"shifts":list_shifts(),"shift_types":list_shift_types(),
            "save_error":"儲存失敗，請稍後再試；原本資料沒有被覆蓋。"
        },status_code=500)
    return RedirectResponse("/quick-schedule?saved=1",303)

@app.post("/schedule/delete")
def schedule_delete(request:Request,work_date:str=Form(...)):
    user,resp=protected(request,"/quick-schedule")
    if resp:return resp
    delete_shift(work_date)
    return RedirectResponse("/quick-schedule?deleted=1",303)

def _excel_text(value):
    if value is None:
        return ""
    if hasattr(value, "strftime") and not isinstance(value, str):
        try:
            return value.strftime("%H:%M")
        except Exception:
            pass
    return str(value).strip()


def _normalise_header(value):
    return re.sub(r"[\s_／/()（）]+", "", _excel_text(value)).lower()


def _normalise_store(value):
    text=_excel_text(value).replace(" ", "")
    if "溪洲" in text or text.upper()=="C":
        return "溪洲"
    if "建昌" in text or text.upper()=="B":
        return "建昌"
    return ""


def _normalise_shift_name(store_value, shift_value):
    shift=_excel_text(shift_value).replace(" ", "")
    if not shift:
        return ""
    if "休" in shift:
        return "休假"
    store=_normalise_store(store_value)
    if "大夜" in shift or shift in {"夜班", "night"}:
        return f"{store}大夜" if store else ""
    if "晚班" in shift or shift in {"晚", "late"}:
        return f"{store}晚班" if store else ""
    # 兼容舊範本直接填入完整班別名稱。
    return shift


def _parse_work_date(raw_date):
    if raw_date is None or raw_date == "":
        return None
    if hasattr(raw_date, "strftime"):
        return raw_date.strftime("%Y-%m-%d")
    text=_excel_text(raw_date).replace(".", "-").replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y%m%d"):
        try:
            return datetime.strptime(text[:19], fmt).date().isoformat()
        except Exception:
            pass
    try:
        return date.fromisoformat(text[:10]).isoformat()
    except Exception:
        return None


def _parse_overtime(value):
    text=_excel_text(value)
    compact=text.replace(" ", "").replace("：", ":")
    if not compact or compact.lower() in {"否", "no", "n", "false", "0", "-", "無"}:
        return False, ""
    match=re.search(r"([01]?\d|2[0-3]):([0-5]\d)", compact)
    if match:
        return True, f"{int(match.group(1)):02d}:{match.group(2)}"
    return compact.lower() in {"是", "yes", "y", "true", "1", "加班"}, ""


def _header_map(ws):
    aliases={
        "date":{"日期", "工作日期", "上班日期"},
        "store":{"店名", "店舖", "店鋪", "門市", "工作店舖"},
        "shift":{"班別", "班次", "班別名稱"},
        "start":{"上班時間", "開始時間", "上班"},
        "end":{"下班時間", "結束時間", "下班"},
        "overtime":{"加班", "加班時間", "加班至"},
        "note":{"備註", "說明"},
    }
    normalised={key:{_normalise_header(v) for v in values} for key,values in aliases.items()}
    for row_no in range(1, min(ws.max_row or 1, 12)+1):
        mapping={}
        for col_no in range(1, (ws.max_column or 1)+1):
            header=_normalise_header(ws.cell(row_no,col_no).value)
            for key,values in normalised.items():
                if header in values:
                    mapping[key]=col_no
        if "date" in mapping and "shift" in mapping:
            return row_no,mapping
    return None,{}


def parse_excel(content:bytes, filename:str=""):
    """讀取 Work Life 班表，支援新版直式表格與舊版月班表範本。"""
    try:
        wb=load_workbook(
            BytesIO(content),
            data_only=True,
            keep_vba=filename.lower().endswith(".xlsm")
        )
    except Exception as exc:
        raise ValueError("Excel 檔案無法開啟，請確認檔案未損壞，且格式為 .xlsx 或 .xlsm。") from exc

    types={str(x["name"]).strip():x for x in list_shift_types()}
    rows=[];errors=[];matched_sheets=0
    seen_dates=set()
    for ws in wb.worksheets:
        header_row,columns=_header_map(ws)
        if header_row:
            matched_sheets+=1
            for row_no in range(header_row+1,(ws.max_row or header_row)+1):
                raw_date=ws.cell(row_no,columns["date"]).value
                raw_shift=ws.cell(row_no,columns["shift"]).value
                raw_store=ws.cell(row_no,columns.get("store",0)).value if columns.get("store") else ""
                if not raw_date and not raw_shift:
                    continue
                work_date=_parse_work_date(raw_date)
                if not work_date:
                    errors.append(f"{ws.title} 第{row_no}列日期無法讀取")
                    continue
                shift_name=_normalise_shift_name(raw_store,raw_shift)
                if not shift_name:
                    errors.append(f"{ws.title} 第{row_no}列缺少店名或班別")
                    continue
                if shift_name not in types:
                    errors.append(f"{ws.title} 第{row_no}列班別不存在：{shift_name}")
                    continue
                overtime,overtime_end=_parse_overtime(ws.cell(row_no,columns.get("overtime",0)).value if columns.get("overtime") else "")
                note=_excel_text(ws.cell(row_no,columns.get("note",0)).value) if columns.get("note") else ""
                if work_date in seen_dates:
                    errors.append(f"{ws.title} 第{row_no}列日期重複：{work_date}")
                    continue
                rows.append({
                    "work_date":work_date,
                    "shift_type_id":types[shift_name]["id"],
                    "overtime":overtime,
                    "overtime_end":overtime_end or ("08:00" if overtime else ""),
                    "note":note
                })
                seen_dates.add(work_date)
            continue

        # 舊版 Work Life 月班表：A 日期、C 班別、D 是否加班、E 加班時間、F 備註。
        if "月班表" not in ws.title:
            continue
        matched_sheets+=1
        for row_no in range(5,(ws.max_row or 4)+1):
            work_date=_parse_work_date(ws.cell(row_no,1).value)
            shift_name=_excel_text(ws.cell(row_no,3).value)
            if not work_date or not shift_name:
                continue
            if shift_name not in types:
                errors.append(f"{ws.title} 第{row_no}列班別不存在：{shift_name}")
                continue
            overtime_text=_excel_text(ws.cell(row_no,4).value)
            overtime,overtime_from_text=_parse_overtime(overtime_text)
            overtime_end=_excel_text(ws.cell(row_no,5).value) or overtime_from_text
            note=_excel_text(ws.cell(row_no,6).value)
            if work_date in seen_dates:
                errors.append(f"{ws.title} 第{row_no}列日期重複：{work_date}")
                continue
            rows.append({
                "work_date":work_date,
                "shift_type_id":types[shift_name]["id"],
                "overtime":overtime,
                "overtime_end":overtime_end or ("08:00" if overtime else ""),
                "note":note
            })
            seen_dates.add(work_date)
    if matched_sheets==0:
        raise ValueError("找不到班表欄位。請確認第一列包含：日期、店名、班別、上班時間、下班時間、加班。")
    return rows,errors

@app.post("/quick-schedule/upload",response_class=HTMLResponse)
async def excel_upload(request:Request,file:UploadFile=File(...),overwrite:str=Form("")):
    user,resp=protected(request,"/quick-schedule")
    if resp:return resp
    context={"request":request,"user":user,"shifts":list_shifts(),"shift_types":list_shift_types()}
    filename=(file.filename or "").strip()
    if not filename.lower().endswith((".xlsx",".xlsm")):
        context["upload_error"]="請上傳 .xlsx 或 .xlsm 格式的 Excel 班表。"
        return templates.TemplateResponse("schedule.html",context,status_code=400)
    content=await file.read()
    if not content:
        context["upload_error"]="上傳的 Excel 檔案是空的，請重新選擇檔案。"
        return templates.TemplateResponse("schedule.html",context,status_code=400)
    if len(content)>10*1024*1024:
        context["upload_error"]="Excel 檔案超過 10MB，請縮小檔案後再上傳。"
        return templates.TemplateResponse("schedule.html",context,status_code=400)
    try:
        rows,errors=parse_excel(content,filename)
    except ValueError as exc:
        context["upload_error"]=str(exc)
        return templates.TemplateResponse("schedule.html",context,status_code=400)
    if not rows:
        context["upload_error"]="班表內沒有可匯入的資料，請先在班別欄選擇班別。"
        if errors:
            context["upload_error"] += " " + "、".join(errors[:8])
        return templates.TemplateResponse("schedule.html",context,status_code=400)

    existing={x["work_date"] for x in list_shifts()}
    imported=0;skipped=0
    for row in rows:
        if row["work_date"] in existing and overwrite!="1":
            skipped+=1
            continue
        save_shift(row["work_date"],row["shift_type_id"],row["overtime"],row["overtime_end"],False,row["note"])
        imported+=1
        existing.add(row["work_date"])
    context.update({"shifts":list_shifts(),"upload_result":{"imported":imported,"skipped":skipped,"errors":errors}})
    return templates.TemplateResponse("schedule.html",context)

@app.get("/work-records",response_class=HTMLResponse)
def work_page(request:Request,work_date:str|None=None):
    user,resp=protected(request,"/work-records")
    if resp:return resp
    selected=work_date or resolve_active_work_date()
    raw_shift=get_shift(selected)
    return templates.TemplateResponse("work_records.html",{
        "request":request,"user":user,"selected_date":selected,
        "shift":decorate_shift_dates(selected,raw_shift),
        "daily_work":decorate_timed_rows(selected,raw_shift,list_daily_work(selected),"scheduled_time"),
        "daily_logistics":decorate_timed_rows(selected,raw_shift,list_daily_logistics(selected),"start_time"),
        "attendance":attendance_for(selected),"logs":list_work_logs()
    })

@app.post("/daily-work/{item_id}/toggle")
def daily_toggle(request:Request,item_id:int,work_date:str=Form(...)):
    user,resp=protected(request,"/work-records")
    if resp:return resp
    toggle_daily_work(item_id)
    return RedirectResponse(f"/work-records?work_date={work_date}&updated=1",303)

@app.post("/attendance/check-in")
def attendance_check_in(request:Request,work_date:str=Form(...)):
    user,resp=protected(request,"/work-records")
    if resp:return resp
    check_in(work_date)
    return RedirectResponse(f"/work-records?work_date={work_date}&updated=1",303)

@app.post("/attendance/check-out")
def attendance_check_out(request:Request,work_date:str=Form(...)):
    user,resp=protected(request,"/work-records")
    if resp:return resp
    check_out(work_date)
    return RedirectResponse(f"/work-records?work_date={work_date}&updated=1",303)

@app.post("/logistics/{item_id}/arrive")
def logistics_start(request:Request,item_id:int,work_date:str=Form(...)):
    user,resp=protected(request,"/work-records")
    if resp:return resp
    logistics_arrive(item_id)
    return RedirectResponse(f"/work-records?work_date={work_date}&updated=1",303)

@app.post("/logistics/{item_id}/complete")
def logistics_finish(request:Request,item_id:int,work_date:str=Form(...)):
    user,resp=protected(request,"/work-records")
    if resp:return resp
    logistics_complete(item_id)
    return RedirectResponse(f"/work-records?work_date={work_date}&updated=1",303)

@app.post("/work-records")
def log_add(request:Request,work_date:str=Form(...),category:str=Form(...),title:str=Form(...),amount:int=Form(0),note:str=Form("")):
    user,resp=protected(request,"/work-records")
    if resp:return resp
    add_work_log(work_date,category,title,max(amount,0),note)
    return RedirectResponse(f"/work-records?work_date={work_date}&updated=1",303)

@app.get("/logistics-settings",response_class=HTMLResponse)
def logistics_settings_page(request:Request):
    user,resp=protected(request,"/logistics-settings")
    if resp:return resp
    return templates.TemplateResponse("logistics.html",{
        "request":request,"user":user,"items":list_logistics_settings()
    })

@app.post("/logistics-settings")
def logistics_settings_save(
    request:Request,item_id:str=Form(""),name:str=Form(...),icon:str=Form("🚚"),
    start_time:str=Form(...),end_time:str=Form(...),content:str=Form(""),
    applies_b:str=Form(""),applies_c:str=Form(""),applies_late:str=Form(""),
    applies_night:str=Form(""),remind_minutes:int=Form(10),
    line_push:str=Form(""),show_carousel:str=Form(""),is_active:str=Form("")
):
    user,resp=protected(request,"/logistics-settings")
    if resp:return resp
    try:
        save_logistics_setting(
            int(item_id) if item_id.isdigit() else None,name.strip(),icon.strip() or "🚚",
            start_time,end_time,content.strip(),applies_b=="1",applies_c=="1",
            applies_late=="1",applies_night=="1",max(remind_minutes,0),
            line_push=="1",show_carousel=="1",is_active=="1"
        )
    except (ValueError, sqlite3.DatabaseError) as exc:
        return templates.TemplateResponse("logistics.html", {
            "request":request,"user":user,"items":list_logistics_settings(),"save_error":str(exc)
        }, status_code=400)
    return RedirectResponse("/logistics-settings?saved=1",303)

@app.post("/logistics-settings/{item_id}/delete")
def logistics_settings_delete(request:Request,item_id:int):
    user,resp=protected(request,"/logistics-settings")
    if resp:return resp
    delete_logistics_setting(item_id)
    return RedirectResponse("/logistics-settings?deleted=1",303)

@app.get("/activities",response_class=HTMLResponse)
def activities_page(request:Request,edit_id:int|None=None):
    user,resp=protected(request,"/activities")
    if resp:return resp
    return templates.TemplateResponse("activities.html",{
        "request":request,"user":user,"today":date.today().isoformat(),
        "activities":list_activities(),"editing":get_activity(edit_id) if edit_id else None
    })

@app.post("/activities")
def activity_add(
    request:Request,title:str=Form(...),activity_date:str=Form(...),
    activity_time:str=Form(""),location:str=Form(""),note:str=Form(""),
    line_push:str=Form(""),remind_minutes:int=Form(30)
):
    user,resp=protected(request,"/activities")
    if resp:return resp
    add_activity(title.strip(),activity_date,activity_time,location.strip(),note.strip(),line_push=="1",max(remind_minutes,0))
    return RedirectResponse("/activities?saved=1",303)

@app.post("/activities/{item_id}/edit")
def activity_edit(
    request:Request,item_id:int,title:str=Form(...),activity_date:str=Form(...),
    activity_time:str=Form(""),location:str=Form(""),note:str=Form(""),
    line_push:str=Form(""),remind_minutes:int=Form(30)
):
    user,resp=protected(request,"/activities")
    if resp:return resp
    update_activity(item_id,title.strip(),activity_date,activity_time,location.strip(),note.strip(),line_push=="1",max(remind_minutes,0))
    return RedirectResponse("/activities?updated=1",303)

@app.post("/activities/{item_id}/complete")
def activity_complete(request:Request,item_id:int):
    user,resp=protected(request,"/activities")
    if resp:return resp
    complete_activity(item_id,True)
    return RedirectResponse("/activities?completed=1",303)

@app.post("/activities/{item_id}/restore")
def activity_restore(request:Request,item_id:int):
    user,resp=protected(request,"/activities")
    if resp:return resp
    complete_activity(item_id,False)
    return RedirectResponse("/activities?restored=1",303)

@app.post("/activities/{item_id}/delete")
def activity_delete(request:Request,item_id:int):
    user,resp=protected(request,"/activities")
    if resp:return resp
    delete_activity(item_id)
    return RedirectResponse("/activities?deleted=1",303)

@app.get("/notifications",response_class=HTMLResponse)
def notification_page(request:Request):
    user,resp=protected(request,"/notifications")
    if resp:return resp
    return templates.TemplateResponse("notifications.html",{
        "request":request,"user":user,"today":date.today().isoformat(),
        "reminders":list_reminders(),"logs":list_notification_logs()
    })

@app.post("/reminders")
def reminder_add(
    request:Request,title:str=Form(...),reminder_type:str=Form("自訂"),
    remind_date:str=Form(...),remind_time:str=Form(...),note:str=Form(""),
    related_url:str=Form("/dashboard"),line_push:str=Form(""),show_carousel:str=Form("")
):
    user,resp=protected(request,"/notifications")
    if resp:return resp
    add_reminder(title.strip(),reminder_type,remind_date,remind_time,note.strip(),
                 related_url,line_push=="1",show_carousel=="1")
    return RedirectResponse("/notifications?saved=1",303)

@app.post("/reminders/{item_id}/complete")
def reminder_done(request:Request,item_id:int):
    user,resp=protected(request,"/notifications")
    if resp:return resp
    complete_reminder(item_id)
    return RedirectResponse("/notifications?completed=1",303)

@app.post("/reminders/{item_id}/delete")
def reminder_remove(request:Request,item_id:int):
    user,resp=protected(request,"/notifications")
    if resp:return resp
    delete_reminder(item_id)
    return RedirectResponse("/notifications?deleted=1",303)

@app.get("/tasks",response_class=HTMLResponse)
def tasks_page(request:Request):
    user,resp=protected(request,"/tasks")
    if resp:return resp
    return templates.TemplateResponse("tasks.html",{"request":request,"user":user,"tasks":list_tasks()})

@app.post("/tasks")
def task_add(request:Request,title:str=Form(...),due_date:str=Form(""),note:str=Form("")):
    user,resp=protected(request,"/tasks")
    if resp:return resp
    add_task(title,due_date,note)
    return RedirectResponse("/tasks?saved=1",303)

@app.post("/tasks/{item_id}/toggle")
def task_toggle(request:Request,item_id:int):
    user,resp=protected(request,"/tasks")
    if resp:return resp
    toggle_task(item_id)
    return RedirectResponse("/tasks?updated=1",303)

@app.post("/tasks/{item_id}/delete")
def task_remove(request:Request,item_id:int):
    user,resp=protected(request,"/tasks")
    if resp:return resp
    delete_task(item_id)
    return RedirectResponse("/tasks?deleted=1",303)

@app.get("/profile",response_class=HTMLResponse)
def profile_page(request:Request):
    user,resp=protected(request,"/profile")
    if resp:return resp
    return templates.TemplateResponse("profile.html",{
        "request":request,"user":user,
        "line_push":get_setting("line_push","1"),
        "line_push_ready":bool(settings.line_channel_access_token and settings.owner_line_user_id)
    })

@app.post("/profile")
def profile_save(request:Request,line_push:str=Form("")):
    user,resp=protected(request,"/profile")
    if resp:return resp
    set_setting("line_push","1" if line_push=="1" else "0")
    return RedirectResponse("/profile?saved=1",303)

@app.post("/profile/test-line-push")
def profile_test_line_push(request:Request):
    user,resp=protected(request,"/profile")
    if resp:return resp
    title="LINE 測試通知"
    now=datetime.now(TAIPEI_TZ)
    content=f"Work Life LINE 推播測試成功。\n測試時間：{now.strftime('%Y/%m/%d %H:%M:%S')}"
    try:
        push_message(
            settings.owner_line_user_id,
            [reminder_flex(title,content,f"{settings.base_url}/profile")]
        )
        add_notification_log(title,content,"成功")
        return RedirectResponse("/profile?test_push=success",303)
    except Exception as exc:
        error=str(exc)[:300]
        add_notification_log(title,error,"失敗")
        return RedirectResponse("/profile?test_push=failed&reason="+urllib.parse.quote(error),303)

@app.get("/api/dashboard-state")
def dashboard_state(request:Request):
    user,resp=protected(request,"/dashboard")
    if resp:return JSONResponse({"ok":False},401)
    now=datetime.now()
    today=date.today().isoformat()
    work_date,shift,daily_work,daily_logistics,next_work_item=dashboard_work_context(now)
    reminders=[x for x in list_reminders(False) if x["remind_date"]==today]
    return {
        "ok":True,"active_work_date":work_date,
        "shift":shift,
        "daily_work":daily_work,
        "daily_logistics":daily_logistics,
        "next_work_item":next_work_item,
        "attendance":attendance_for(work_date),
        "counts":dashboard_counts(work_date),
        "activities":upcoming_activities(today,5),
        "next_shift":get_next_shift(today),
        "reminder":reminders[0] if reminders else None,
        "updated_at":datetime.now().strftime("%H:%M:%S")
    }

@app.post("/webhook")
async def webhook(request:Request):
    body=(await request.body()).decode("utf-8")
    signature=request.headers.get("X-Line-Signature","")
    if not verify_signature(body,signature):
        return JSONResponse({"ok":False},400)
    data=await request.json()
    for event in data.get("events",[]):
        if event.get("type")=="message" and event.get("message",{}).get("type")=="text":
            text=event["message"]["text"].strip().lower()
            if text in {"/work","work","工作小幫手","班表"}:
                reply_message(event["replyToken"],[work_entry_flex(settings.base_url)])
    return {"ok":True}
